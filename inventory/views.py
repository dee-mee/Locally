from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q, Sum, F, Count, Case, When, Value, IntegerField, DecimalField
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models.functions import Coalesce
import json
from datetime import timedelta

from .models import Product, Category, StockMovement
from .forms import ProductForm, CategoryForm, StockAdjustmentForm, ProductImportForm

# Create your views here.

class ProductBulkActionView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Handle bulk actions for products"""
    permission_required = 'inventory.change_product'
    
    def post(self, request, *args, **kwargs):
        # Get selected product IDs
        product_ids = request.POST.get('product_ids', '').split(',')
        action = request.POST.get('action')
        
        if not product_ids or not product_ids[0]:
            messages.error(request, 'No products selected.')
            return HttpResponseRedirect(reverse('inventory:product_list'))
            
        # Get the products
        products = Product.objects.filter(id__in=product_ids)
        
        if not products.exists():
            messages.error(request, 'No valid products found.')
            return HttpResponseRedirect(reverse('inventory:product_list'))
        
        success_count = 0
        
        # Perform the requested action
        try:
            if action == 'activate':
                success_count = products.update(status='active')
                messages.success(request, f'Successfully activated {success_count} product(s).')
                
            elif action == 'deactivate':
                success_count = products.update(status='inactive')
                messages.success(request, f'Successfully deactivated {success_count} product(s).')
                
            elif action in ['add_to_category', 'remove_from_category']:
                category_id = request.POST.get('category')
                if not category_id:
                    messages.error(request, 'Category is required for this action.')
                    return HttpResponseRedirect(reverse('inventory:product_list'))
                    
                try:
                    category = Category.objects.get(id=category_id)
                    
                    for product in products:
                        if action == 'add_to_category':
                            product.categories.add(category)
                        else:
                            product.categories.remove(category)
                        success_count += 1
                        
                    messages.success(request, f'Successfully updated categories for {success_count} product(s).')
                    
                except Category.DoesNotExist:
                    messages.error(request, 'Invalid category selected.')
                    return HttpResponseRedirect(reverse('inventory:product_list'))
                
            elif action == 'delete':
                if not request.user.has_perm('inventory.delete_product'):
                    messages.error(request, 'You do not have permission to delete products.')
                    return HttpResponseRedirect(reverse('inventory:product_list'))
                    
                success_count = products.count()
                products.delete()
                messages.success(request, f'Successfully deleted {success_count} product(s).')
                
            else:
                messages.error(request, 'Invalid action specified.')
                
        except Exception as e:
            messages.error(request, f'Error performing bulk action: {str(e)}')
        
        return HttpResponseRedirect(reverse('inventory:product_list'))
class CategoryListView(LoginRequiredMixin, ListView):
    """List all product categories"""
    model = Category
    template_name = 'inventory/category_list.html'
    context_object_name = 'categories'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Category.objects.all().annotate(
            product_count=Count('products', distinct=True)
        ).order_by('name')
        
        # Search functionality
        search_query = self.request.GET.get('q')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        return queryset


class CategoryDetailView(LoginRequiredMixin, DetailView):
    """View details of a specific category"""
    model = Category
    template_name = 'inventory/category_detail.html'
    context_object_name = 'category'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        category = self.get_object()
        
        # Get products in this category with stock information
        products = category.products.all().annotate(
            total_cost=Coalesce(
                F('stock_quantity') * F('cost_price'),
                Value(0, output_field=DecimalField())
            ),
            total_value=Coalesce(
                F('stock_quantity') * F('selling_price'),
                Value(0, output_field=DecimalField())
            )
        ).order_by('name')
        
        # Pagination
        paginator = Paginator(products, 10)
        page = self.request.GET.get('page')
        
        try:
            products = paginator.page(page)
        except PageNotAnInteger:
            products = paginator.page(1)
        except EmptyPage:
            products = paginator.page(paginator.num_pages)
        
        # Calculate category statistics
        stats = {
            'total_products': products.paginator.count,
            'total_items': products.object_list.aggregate(
                total=Coalesce(Sum('stock_quantity'), Value(0, output_field=DecimalField()))
            )['total'],
            'total_cost': products.object_list.aggregate(
                total=Coalesce(Sum('total_cost'), Value(0, output_field=DecimalField()))
            )['total'],
            'total_value': products.object_list.aggregate(
                total=Coalesce(Sum('total_value'), Value(0, output_field=DecimalField()))
            )['total'],
        }
        
        context.update({
            'products': products,
            'stats': stats,
        })
        return context


class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new product category"""
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/category_form.html'
    permission_required = 'inventory.add_category'
    
    def get_success_url(self):
        messages.success(self.request, 'Category created successfully.')
        return reverse_lazy('inventory:category_detail', kwargs={'slug': self.object.slug})
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing product category"""
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/category_form.html'
    permission_required = 'inventory.change_category'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    
    def get_success_url(self):
        messages.success(self.request, 'Category updated successfully.')
        return reverse_lazy('inventory:category_detail', kwargs={'slug': self.object.slug})


class CategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a product category"""
    model = Category
    template_name = 'inventory/category_confirm_delete.html'
    permission_required = 'inventory.delete_category'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('inventory:category_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Category deleted successfully.')
        return super().delete(request, *args, **kwargs)


# Product Views
class ProductListView(LoginRequiredMixin, ListView):
    """List all products with filtering and search"""
    model = Product
    template_name = 'inventory/product_list.html'
    context_object_name = 'products'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Product.objects.select_related('category').annotate(
            total_value=Coalesce(
                F('stock_quantity') * F('selling_price'),
                Value(0, output_field=DecimalField())
            )
        ).order_by('name')
        
        # Apply filters
        category = self.request.GET.get('category')
        status = self.request.GET.get('status')
        stock_status = self.request.GET.get('stock_status')
        search_query = self.request.GET.get('q')
        
        if category:
            queryset = queryset.filter(category__slug=category)
            
        if status:
            queryset = queryset.filter(status=status)
            
        if stock_status:
            if stock_status == 'low':
                queryset = queryset.filter(stock_quantity__lte=F('reorder_level'))
            elif stock_status == 'out':
                queryset = queryset.filter(stock_quantity__lte=0)
            elif stock_status == 'in_stock':
                queryset = queryset.filter(stock_quantity__gt=F('reorder_level'))
                
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(sku__icontains=search_query) |
                Q(barcode__icontains=search_query) |
                Q(description__icontains=search_query)
            )
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add categories for filter dropdown
        context['categories'] = Category.objects.all()
        
        # Add filter values to maintain state
        context['current_category'] = self.request.GET.get('category', '')
        context['current_status'] = self.request.GET.get('status', '')
        context['current_stock_status'] = self.request.GET.get('stock_status', '')
        context['search_query'] = self.request.GET.get('q', '')
        
        # Calculate inventory summary
        if context['products']:
            context['total_products'] = context['products'].paginator.count
            context['total_items'] = sum(p.stock_quantity for p in context['products'])
            context['total_value'] = sum(
                p.stock_quantity * p.selling_price 
                for p in context['products'] 
                if p.stock_quantity and p.selling_price
            )
        
        return context


class ProductDetailView(LoginRequiredMixin, DetailView):
    """View details of a specific product"""
    model = Product
    template_name = 'inventory/product_detail.html'
    context_object_name = 'product'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = self.get_object()
        
        # Get stock movements
        movements = product.stock_movements.all().order_by('-created_at')
        
        # Paginate movements
        paginator = Paginator(movements, 10)
        page = self.request.GET.get('page')
        
        try:
            movements = paginator.page(page)
        except PageNotAnInteger:
            movements = paginator.page(1)
        except EmptyPage:
            movements = paginator.page(paginator.num_pages)
        
        # Calculate statistics
        stats = {
            'total_sold': product.stock_movements.filter(
                movement_type='sale'
            ).aggregate(total=Sum('quantity'))['total'] or 0,
            'total_purchased': product.stock_movements.filter(
                movement_type='purchase'
            ).aggregate(total=Sum('quantity'))['total'] or 0,
            'total_adjusted': product.stock_movements.filter(
                movement_type='adjustment'
            ).aggregate(total=Sum('quantity'))['total'] or 0,
        }
        
        context.update({
            'movements': movements,
            'stats': stats,
            'adjustment_form': StockAdjustmentForm(initial={
                'product': product,
                'movement_type': 'adjustment',
                'quantity': 0,
            }),
        })
        return context


class ProductCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new product"""
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    permission_required = 'inventory.add_product'
    
    def get_success_url(self):
        messages.success(self.request, 'Product created successfully.')
        return reverse_lazy('inventory:product_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class ProductUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing product"""
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    permission_required = 'inventory.change_product'
    
    def get_success_url(self):
        messages.success(self.request, 'Product updated successfully.')
        return reverse_lazy('inventory:product_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.last_updated_by = self.request.user
        return super().form_valid(form)


class ProductDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a product"""
    model = Product
    template_name = 'inventory/product_confirm_delete.html'
    permission_required = 'inventory.delete_product'
    success_url = reverse_lazy('inventory:product_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Product deleted successfully.')
        return super().delete(request, *args, **kwargs)


# Stock Management Views
@login_required
@permission_required('inventory.change_product')
def adjust_stock(request, product_id):
    """Adjust product stock quantity"""
    product = get_object_or_404(Product, pk=product_id)
    
    if request.method == 'POST':
        form = StockAdjustmentForm(request.POST)
        if form.is_valid():
            movement = form.save(commit=False)
            movement.product = product
            movement.created_by = request.user
            movement.save()
            
            messages.success(request, 'Stock updated successfully.')
            return redirect('inventory:product_detail', pk=product.pk)
    else:
        form = StockAdjustmentForm(initial={
            'product': product,
            'movement_type': 'adjustment',
            'quantity': 0,
        })
    
    return render(request, 'inventory/stock_adjustment.html', {
        'form': form,
        'product': product,
    })


@login_required
@permission_required('inventory.add_product')
def import_products(request):
    """Import products from CSV/Excel using django-import-export"""
    from tablib import Dataset
    from .resources import ProductResource
    import pandas as pd
    
    import_results = None
    
    if request.method == 'POST':
        try:
            # Get the uploaded file
            file = request.FILES.get('file')
            if not file:
                messages.error(request, 'Please select a file to upload')
                return redirect('inventory:import_products')
            
            # Validate file extension
            if not file.name.endswith(('.xlsx', '.xls', '.csv')):
                messages.error(request, 'Invalid file type. Please upload a .xlsx, .xls, or .csv file')
                return redirect('inventory:import_products')
            
            # Read the file based on its type
            if file.name.endswith('.csv'):
                # For CSV files
                dataset = Dataset()
                dataset.load(file.read().decode('utf-8'), format='csv')
            else:
                # For Excel files
                df = pd.read_excel(file)
                dataset = Dataset()
                dataset.df = df
            
            # Check if the file has headers
            has_headers = request.POST.get('has_headers') == 'on'
            update_existing = request.POST.get('update_existing') == 'on'
            
            # Import the data
            product_resource = ProductResource()
            
            if has_headers:
                # If file has headers, let import_data handle column mapping
                result = product_resource.import_data(
                    dataset,
                    dry_run=False,
                    raise_errors=False,
                    use_transactions=True,
                    collect_failed_rows=True
                )
            else:
                # If no headers, use the first row as data
                result = product_resource.import_data(
                    dataset,
                    dry_run=False,
                    raise_errors=False,
                    use_transactions=True,
                    collect_failed_rows=True
                )
            
            # Prepare import results
            import_results = {
                'success': not result.has_errors(),
                'imported': result.totals.get('new', 0),
                'updated': result.totals.get('update', 0),
                'skipped': result.totals.get('skip', 0),
                'errors': []
            }
            
            # Collect any errors
            if result.has_errors():
                for row in result.invalid_rows:
                    import_results['errors'].append({
                        'row': row.number,
                        'message': row.error.message if hasattr(row.error, 'message') else str(row.error)
                    })
            
            if import_results['success']:
                messages.success(
                    request,
                    f'Successfully imported {import_results["imported"]} products. ' \
                    f'Updated {import_results["updated"]} existing products.'
                )
            else:
                messages.warning(
                    request,
                    f'Import completed with {len(import_results["errors"])} errors. ' \
                    f'Imported: {import_results["imported"]}, Updated: {import_results["updated"]}.'
                )
            
        except Exception as e:
            import traceback
            error_msg = str(e)
            import_results = {
                'success': False,
                'error': error_msg,
                'traceback': traceback.format_exc()
            }
            messages.error(request, f'Error importing products: {error_msg}')
    
    return render(request, 'inventory/import_products.html', {
        'import_results': import_results
    })


# AJAX Views
@login_required
@require_POST
def ajax_product_search(request):
    """AJAX endpoint for product search"""
    query = request.POST.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({'results': []})
    
    products = Product.objects.filter(
        Q(name__icontains=query) |
        Q(sku__iexact=query) |
        Q(barcode__iexact=query)
    ).filter(status=Product.ACTIVE).order_by('name')[:10]
    
    results = [{
        'id': p.id,
        'name': p.name,
        'sku': p.sku,
        'barcode': p.barcode,
        'price': str(p.selling_price),
        'stock': float(p.stock_quantity),
        'unit': p.get_unit_display(),
    } for p in products]
    
    return JsonResponse({'results': results})


@login_required
@require_POST
def ajax_get_product(request, product_id):
    """AJAX endpoint to get product details"""
    try:
        product = Product.objects.get(pk=product_id, status=Product.ACTIVE)
        return JsonResponse({
            'success': True,
            'product': {
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'barcode': product.barcode,
                'price': str(product.selling_price),
                'stock': float(product.stock_quantity),
                'unit': product.get_unit_display(),
                'tax_rate': float(product.tax_rate),
            }
        })
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# Reports
@login_required
def inventory_report(request):
    """Generate inventory report"""
    products = Product.objects.select_related('category').annotate(
        total_cost=Coalesce(
            F('stock_quantity') * F('cost_price'),
            Value(0, output_field=DecimalField())
        ),
        total_value=Coalesce(
            F('stock_quantity') * F('selling_price'),
            Value(0, output_field=DecimalField())
        )
    ).order_by('category__name', 'name')
    
    # Apply filters
    category = request.GET.get('category')
    status = request.GET.get('status')
    stock_status = request.GET.get('stock_status')
    
    if category:
        products = products.filter(category__slug=category)
        
    if status:
        products = products.filter(status=status)
        
    if stock_status == 'low':
        products = products.filter(stock_quantity__lte=F('reorder_level'))
    elif stock_status == 'out':
        products = products.filter(stock_quantity__lte=0)
    elif stock_status == 'in_stock':
        products = products.filter(stock_quantity__gt=F('reorder_level'))
    
    # Calculate summary
    summary = {
        'total_products': products.count(),
        'total_items': products.aggregate(
            total=Coalesce(Sum('stock_quantity'), Value(0, output_field=DecimalField()))
        )['total'],
        'total_cost': products.aggregate(
            total=Coalesce(Sum('total_cost'), Value(0, output_field=DecimalField()))
        )['total'],
        'total_value': products.aggregate(
            total=Coalesce(Sum('total_value'), Value(0, output_field=DecimalField()))
        )['total'],
    }
    
    # Export to CSV/Excel if requested
    export_format = request.GET.get('export')
    if export_format in ['csv', 'excel']:
        response = HttpResponse(content_type=
            'text/csv' if export_format == 'csv' 
            else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="inventory_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.{export_format}"'
        
        if export_format == 'csv':
            writer = csv.writer(response)
            # Write headers
            writer.writerow([
                'SKU', 'Name', 'Category', 'Status', 'Stock', 'Unit', 
                'Cost Price', 'Selling Price', 'Stock Value', 'Reorder Level'
            ])
            
            # Write data rows
            for product in products:
                writer.writerow([
                    product.sku,
                    product.name,
                    str(product.category),
                    product.get_status_display(),
                    product.stock_quantity,
                    product.get_unit_display(),
                    product.cost_price,
                    product.selling_price,
                    product.total_value,
                    product.reorder_level,
                ])
        else:
            # Excel export using openpyxl
            from openpyxl import Workbook
            from openpyxl.writer.excel import save_virtual_workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventory Report"
            
            # Write headers
            headers = [
                'SKU', 'Name', 'Category', 'Status', 'Stock', 'Unit', 
                'Cost Price', 'Selling Price', 'Stock Value', 'Reorder Level', 'Last Updated'
            ]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Write data rows
            for row_num, product in enumerate(products, 2):
                ws.cell(row=row_num, column=1, value=product.sku)
                ws.cell(row=row_num, column=2, value=product.name)
                ws.cell(row=row_num, column=3, value=str(product.category))
                ws.cell(row=row_num, column=4, value=product.get_status_display())
                ws.cell(row=row_num, column=5, value=float(product.stock_quantity))
                ws.cell(row=row_num, column=6, value=product.get_unit_display())
                ws.cell(row=row_num, column=7, value=float(product.cost_price))
                ws.cell(row=row_num, column=8, value=float(product.selling_price))
                ws.cell(row=row_num, column=9, value=float(product.total_value))
                ws.cell(row=row_num, column=10, value=float(product.reorder_level))
                ws.cell(row=row_num, column=11, value=product.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook to the response
            response.write(save_virtual_workbook(wb))
        
        return response
    
    # For HTML view
    categories = Category.objects.all().order_by('name')
    
    return render(request, 'inventory/reports/inventory.html', {
        'products': products,
        'categories': categories,
        'summary': summary,
        'current_category': category,
        'current_status': status,
        'current_stock_status': stock_status,
    })


@login_required
def stock_movements_report(request):
    """Generate stock movements report"""
    movements = StockMovement.objects.select_related('product', 'created_by')
    
    # Apply filters
    product_id = request.GET.get('product')
    movement_type = request.GET.get('movement_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if product_id:
        movements = movements.filter(product_id=product_id)
    
    if movement_type:
        movements = movements.filter(movement_type=movement_type)
    
    if date_from:
        movements = movements.filter(created_at__date__gte=date_from)
    
    if date_to:
        movements = movements.filter(created_at__date__lte=date_to)
    
    # Order by most recent first
    movements = movements.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(movements, 50)
    page = request.GET.get('page')
    
    try:
        movements = paginator.page(page)
    except PageNotAnInteger:
        movements = paginator.page(1)
    except EmptyPage:
        movements = paginator.page(paginator.num_pages)
    
    # Get filter options
    products = Product.objects.filter(
        id__in=movements.object_list.values_list('product_id', flat=True).distinct()
    ).order_by('name')
    
    return render(request, 'inventory/reports/stock_movements.html', {
        'movements': movements,
        'products': products,
        'movement_types': dict(StockMovement.MOVEMENT_TYPES),
        'filters': {
            'product': product_id,
            'movement_type': movement_type,
            'date_from': date_from,
            'date_to': date_to,
        },
    })
